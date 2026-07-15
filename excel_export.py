#!/usr/bin/env python3
"""
Excel Export Module for Gaybeck Starkids SMS
Provides comprehensive Excel export functionality for all database tables
Supports: Student records, Class data, Financial reports, Teacher directory
"""

import sqlite3
import os
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class ExcelExporter:
    """Export database tables to professional Excel files"""
    
    def __init__(self, db_path):
        """Initialize Excel exporter with database path"""
        self.db_path = db_path
        self.conn = None
        self.cursor = None
        
    def connect(self):
        """Connect to database"""
        self.conn = sqlite3.connect(self.db_path)
        self.cursor = self.conn.cursor()
        
    def disconnect(self):
        """Close database connection"""
        if self.conn:
            self.conn.close()
    
    def _apply_header_style(self, worksheet, header_row=1):
        """Apply professional styling to header row"""
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        for cell in worksheet[header_row]:
            if cell.value:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
    
    def _apply_data_style(self, worksheet, start_row=2, end_row=None):
        """Apply styling to data rows"""
        data_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        if end_row is None:
            end_row = worksheet.max_row
        
        for row in worksheet.iter_rows(min_row=start_row, max_row=end_row):
            for cell in row:
                cell.alignment = data_alignment
                cell.border = border
    
    def _auto_adjust_columns(self, worksheet):
        """Auto-adjust column widths"""
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def export_students_full(self, output_path=None):
        """Export all students to Excel with all details"""
        if not OPENPYXL_AVAILABLE:
            return False, "openpyxl not installed. Install with: pip install openpyxl"
        
        try:
            self.connect()
            
            # Get students data with class names
            self.cursor.execute('''
                SELECT 
                    s.student_id,
                    s.name,
                    s.date_of_birth,
                    s.gender,
                    s.phone,
                    s.parent_email,
                    s.father_name,
                    s.mother_name,
                    s.address,
                    c.class_name,
                    s.bus_fee,
                    s.monthly_fee,
                    s.date_of_admission
                FROM students s
                LEFT JOIN classes c ON s.class_id = c.id
                ORDER BY c.class_name, s.name
            ''')
            
            students = self.cursor.fetchall()
            
            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Students"
            
            # Add headers
            headers = [
                'Student ID', 'Name', 'Date of Birth', 'Gender', 'Phone', 'Parent Email',
                'Father Name', 'Mother Name', 'Address', 'Class', 'Bus Fee (GHS)', 'Monthly Fee (GHS)', 'Admission Date'
            ]
            ws.append(headers)
            
            # Add data
            for student in students:
                ws.append(student)
            
            # Apply styling
            self._apply_header_style(ws)
            self._apply_data_style(ws)
            self._auto_adjust_columns(ws)
            
            # Set print options
            ws.page_setup.paperSize = ws.PAPERSIZE_A4
            ws.page_margins.left = 0.5
            ws.page_margins.right = 0.5
            ws.page_margins.top = 0.75
            ws.page_margins.bottom = 0.75
            ws.print_options.horizontalCentered = True
            
            # Save file
            if not output_path:
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                output_path = f'student_database_{timestamp}.xlsx'
            
            wb.save(output_path)
            self.disconnect()
            
            return True, f"✅ Exported {len(students)} students to {output_path}"
            
        except Exception as e:
            self.disconnect()
            return False, f"❌ Error exporting students: {str(e)}"
    
    def export_students_by_class(self, output_path=None):
        """Export students grouped by class with class summaries"""
        if not OPENPYXL_AVAILABLE:
            return False, "openpyxl not installed. Install with: pip install openpyxl"
        
        try:
            self.connect()
            
            # Get unique classes
            self.cursor.execute('SELECT id, class_name FROM classes ORDER BY class_name')
            classes = self.cursor.fetchall()
            
            wb = Workbook()
            wb.remove(wb.active)  # Remove default sheet
            
            total_students = 0
            
            for class_id, class_name in classes:
                # Get students in class
                self.cursor.execute('''
                    SELECT 
                        student_id,
                        name,
                        date_of_birth,
                        gender,
                        phone,
                        monthly_fee,
                        bus_fee
                    FROM students
                    WHERE class_id = ?
                    ORDER BY name
                ''', (class_id,))
                
                students = self.cursor.fetchall()
                
                # Create sheet for class
                ws = wb.create_sheet(title=class_name[:31])  # Excel limit: 31 chars
                
                # Add class summary header
                ws['A1'] = f"Class: {class_name}"
                ws['A1'].font = Font(bold=True, size=14, color="FFFFFF")
                ws['A1'].fill = PatternFill(start_color="34495e", end_color="34495e", fill_type="solid")
                ws.merge_cells('A1:G1')
                
                summary_row = 2
                ws[f'A{summary_row}'] = f"Total Students: {len(students)}"
                ws[f'A{summary_row}'].font = Font(bold=True, size=10)
                
                # Add headers
                headers = ['Student ID', 'Name', 'Date of Birth', 'Gender', 'Phone', 'Monthly Fee (GHS)', 'Bus Fee (GHS)']
                ws.append([''] * 7)  # Empty row for spacing
                ws.append(headers)
                
                # Add data
                for student in students:
                    ws.append(student)
                
                # Apply styling
                self._apply_header_style(ws, header_row=4)
                self._apply_data_style(ws, start_row=5)
                self._auto_adjust_columns(ws)
                
                total_students += len(students)
            
            # Save file
            if not output_path:
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                output_path = f'students_by_class_{timestamp}.xlsx'
            
            wb.save(output_path)
            self.disconnect()
            
            return True, f"✅ Exported {total_students} students in {len(classes)} classes to {output_path}"
            
        except Exception as e:
            self.disconnect()
            return False, f"❌ Error exporting by class: {str(e)}"
    
    def export_student_profile(self, student_id, output_path=None):
        """Export detailed profile for single student"""
        if not OPENPYXL_AVAILABLE:
            return False, "openpyxl not installed. Install with: pip install openpyxl"
        
        try:
            self.connect()
            
            # Get student info
            self.cursor.execute('''
                SELECT * FROM students WHERE student_id = ?
            ''', (student_id,))
            
            student = self.cursor.fetchone()
            if not student:
                return False, f"Student {student_id} not found"
            
            # Get student columns
            self.cursor.execute('PRAGMA table_info(students)')
            columns = [row[1] for row in self.cursor.fetchall()]
            
            # Get grades
            self.cursor.execute('''
                SELECT assignment_name, subject, grade, max_grade, comments
                FROM grades
                WHERE student_id = ?
                ORDER BY date_assigned DESC
                LIMIT 50
            ''', (student_id,))
            grades = self.cursor.fetchall()
            
            # Get attendance
            self.cursor.execute('''
                SELECT date, CASE WHEN present=1 THEN 'Present' ELSE 'Absent' END as status
                FROM attendance
                WHERE student_id = (SELECT id FROM students WHERE student_id = ?)
                ORDER BY date DESC
                LIMIT 30
            ''', (student_id,))
            attendance = self.cursor.fetchall()
            
            # Get fees
            self.cursor.execute('''
                SELECT fee_type, amount_paid, payment_date, amount_due
                FROM fees
                WHERE student_id = (SELECT id FROM students WHERE student_id = ?)
                ORDER BY payment_date DESC
            ''', (student_id,))
            fees = self.cursor.fetchall()
            
            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Student Profile"
            
            # Student Basic Info
            ws['A1'] = "STUDENT PROFILE"
            ws['A1'].font = Font(bold=True, size=16, color="FFFFFF")
            ws['A1'].fill = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
            ws.merge_cells('A1:D1')
            
            row = 3
            ws[f'A{row}'] = "Name:"
            ws[f'B{row}'] = student[columns.index('name')] if 'name' in columns else ''
            row += 1
            
            ws[f'A{row}'] = "Student ID:"
            ws[f'B{row}'] = student_id
            row += 1
            
            # Get class name
            class_id = student[columns.index('class_id')] if 'class_id' in columns else None
            class_name = ''
            if class_id:
                self.cursor.execute('SELECT class_name FROM classes WHERE id = ?', (class_id,))
                result = self.cursor.fetchone()
                class_name = result[0] if result else 'Unknown'
            
            ws[f'A{row}'] = "Class:"
            ws[f'B{row}'] = class_name
            row += 1
            
            ws[f'A{row}'] = "Date of Birth:"
            ws[f'B{row}'] = student[columns.index('date_of_birth')] if 'date_of_birth' in columns else ''
            row += 1
            
            ws[f'A{row}'] = "Gender:"
            ws[f'B{row}'] = student[columns.index('gender')] if 'gender' in columns else ''
            row += 1
            
            ws[f'A{row}'] = "Phone:"
            ws[f'B{row}'] = student[columns.index('phone')] if 'phone' in columns else ''
            row += 1
            
            ws[f'A{row}'] = "Parent Email:"
            ws[f'B{row}'] = student[columns.index('parent_email')] if 'parent_email' in columns else ''
            row += 2
            
            # Grades Section
            ws[f'A{row}'] = "GRADES"
            ws[f'A{row}'].font = Font(bold=True, size=12, color="FFFFFF")
            ws[f'A{row}'].fill = PatternFill(start_color="34495e", end_color="34495e", fill_type="solid")
            row += 1
            
            if grades:
                ws[f'A{row}'] = "Assignment"
                ws[f'B{row}'] = "Subject"
                ws[f'C{row}'] = "Grade"
                ws[f'D{row}'] = "Max Grade"
                ws[f'E{row}'] = "Comments"
                
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="95a5a6", end_color="95a5a6", fill_type="solid")
                for cell in [ws[f'A{row}'], ws[f'B{row}'], ws[f'C{row}'], ws[f'D{row}'], ws[f'E{row}']]:
                    cell.font = header_font
                    cell.fill = header_fill
                
                row += 1
                for grade in grades:
                    ws[f'A{row}'] = grade[0]
                    ws[f'B{row}'] = grade[1]
                    ws[f'C{row}'] = grade[2]
                    ws[f'D{row}'] = grade[3]
                    ws[f'E{row}'] = grade[4]
                    row += 1
            else:
                ws[f'A{row}'] = "No grades recorded"
                row += 1
            
            row += 1
            
            # Attendance Section
            ws[f'A{row}'] = "RECENT ATTENDANCE (Last 30 days)"
            ws[f'A{row}'].font = Font(bold=True, size=12, color="FFFFFF")
            ws[f'A{row}'].fill = PatternFill(start_color="34495e", end_color="34495e", fill_type="solid")
            row += 1
            
            if attendance:
                ws[f'A{row}'] = "Date"
                ws[f'B{row}'] = "Status"
                
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="95a5a6", end_color="95a5a6", fill_type="solid")
                for cell in [ws[f'A{row}'], ws[f'B{row}']]:
                    cell.font = header_font
                    cell.fill = header_fill
                
                row += 1
                for att in attendance[:30]:  # Show last 30
                    ws[f'A{row}'] = att[0]
                    ws[f'B{row}'] = att[1]
                    row += 1
            else:
                ws[f'A{row}'] = "No attendance records"
                row += 1
            
            row += 1
            
            # Fees Section
            ws[f'A{row}'] = "FEES STATUS"
            ws[f'A{row}'].font = Font(bold=True, size=12, color="FFFFFF")
            ws[f'A{row}'].fill = PatternFill(start_color="34495e", end_color="34495e", fill_type="solid")
            row += 1
            
            if fees:
                ws[f'A{row}'] = "Fee Type"
                ws[f'B{row}'] = "Amount Paid (GHS)"
                ws[f'C{row}'] = "Payment Date"
                ws[f'D{row}'] = "Amount Due (GHS)"
                
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="95a5a6", end_color="95a5a6", fill_type="solid")
                for cell in [ws[f'A{row}'], ws[f'B{row}'], ws[f'C{row}'], ws[f'D{row}']]:
                    cell.font = header_font
                    cell.fill = header_fill
                
                row += 1
                for fee in fees:
                    ws[f'A{row}'] = fee[0]
                    ws[f'B{row}'] = fee[1]
                    ws[f'C{row}'] = fee[2]
                    ws[f'D{row}'] = fee[3]
                    row += 1
            else:
                ws[f'A{row}'] = "No fees recorded"
                row += 1
            
            # Auto-adjust columns
            self._auto_adjust_columns(ws)
            
            # Save file
            if not output_path:
                student_name = student[columns.index('name')] if 'name' in columns else 'Unknown'
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                output_path = f'student_profile_{student_id}_{timestamp}.xlsx'
            
            wb.save(output_path)
            self.disconnect()
            
            return True, f"✅ Exported student profile to {output_path}"
            
        except Exception as e:
            self.disconnect()
            return False, f"❌ Error exporting profile: {str(e)}"
    
    def export_teachers_directory(self, output_path=None):
        """Export all teachers to Excel"""
        if not OPENPYXL_AVAILABLE:
            return False, "openpyxl not installed. Install with: pip install openpyxl"
        
        try:
            self.connect()
            
            # Get teachers data with class names
            self.cursor.execute('''
                SELECT 
                    t.id,
                    t.name,
                    c.class_name,
                    t.phone,
                    t.email,
                    t.hire_date,
                    t.starting_salary,
                    t.qualifications
                FROM teachers t
                LEFT JOIN classes c ON t.class_id = c.id
                ORDER BY t.name
            ''')
            
            teachers = self.cursor.fetchall()
            
            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Teachers"
            
            # Add headers
            headers = ['Teacher ID', 'Name', 'Class', 'Phone', 'Email', 'Hire Date', 'Salary (GHS)', 'Qualifications']
            ws.append(headers)
            
            # Add data
            for teacher in teachers:
                ws.append(teacher)
            
            # Apply styling
            self._apply_header_style(ws)
            self._apply_data_style(ws)
            self._auto_adjust_columns(ws)
            
            # Set print options
            ws.page_setup.paperSize = ws.PAPERSIZE_A4
            ws.page_margins.left = 0.5
            ws.page_margins.right = 0.5
            
            # Save file
            if not output_path:
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                output_path = f'teachers_directory_{timestamp}.xlsx'
            
            wb.save(output_path)
            self.disconnect()
            
            return True, f"✅ Exported {len(teachers)} teachers to {output_path}"
            
        except Exception as e:
            self.disconnect()
            return False, f"❌ Error exporting teachers: {str(e)}"
    
    def export_financial_summary(self, output_path=None):
        """Export financial summary by category"""
        if not OPENPYXL_AVAILABLE:
            return False, "openpyxl not installed. Install with: pip install openpyxl"
        
        try:
            self.connect()
            
            # Get income by category
            self.cursor.execute('''
                SELECT 
                    fc.category_name,
                    SUM(ft.amount) as total,
                    COUNT(*) as count
                FROM financial_transactions ft
                JOIN financial_categories fc ON ft.category_id = fc.id
                WHERE ft.transaction_type = 'income'
                GROUP BY fc.category_name
                ORDER BY total DESC
            ''')
            
            income_data = self.cursor.fetchall()
            
            # Get expenses by category
            self.cursor.execute('''
                SELECT 
                    fc.category_name,
                    SUM(ft.amount) as total,
                    COUNT(*) as count
                FROM financial_transactions ft
                JOIN financial_categories fc ON ft.category_id = fc.id
                WHERE ft.transaction_type = 'expense'
                GROUP BY fc.category_name
                ORDER BY total DESC
            ''')
            
            expense_data = self.cursor.fetchall()
            
            # Create workbook
            wb = Workbook()
            
            # Income sheet
            ws_income = wb.active
            ws_income.title = "Income"
            
            ws_income['A1'] = "INCOME SUMMARY"
            ws_income['A1'].font = Font(bold=True, size=14, color="FFFFFF")
            ws_income['A1'].fill = PatternFill(start_color="27ae60", end_color="27ae60", fill_type="solid")
            ws_income.merge_cells('A1:C1')
            
            headers = ['Category', 'Total (GHS)', 'Count']
            ws_income.append([])
            ws_income.append(headers)
            
            for row_data in income_data:
                ws_income.append(row_data)
            
            self._apply_header_style(ws_income, header_row=3)
            self._apply_data_style(ws_income, start_row=4)
            self._auto_adjust_columns(ws_income)
            
            # Add total
            total_income = sum([row[1] for row in income_data])
            last_row = len(income_data) + 4
            ws_income[f'A{last_row}'] = "TOTAL INCOME"
            ws_income[f'A{last_row}'].font = Font(bold=True)
            ws_income[f'B{last_row}'] = total_income
            ws_income[f'B{last_row}'].font = Font(bold=True)
            
            # Expense sheet
            ws_expense = wb.create_sheet("Expenses")
            
            ws_expense['A1'] = "EXPENSE SUMMARY"
            ws_expense['A1'].font = Font(bold=True, size=14, color="FFFFFF")
            ws_expense['A1'].fill = PatternFill(start_color="e74c3c", end_color="e74c3c", fill_type="solid")
            ws_expense.merge_cells('A1:C1')
            
            ws_expense.append([])
            ws_expense.append(headers)
            
            for row_data in expense_data:
                ws_expense.append(row_data)
            
            self._apply_header_style(ws_expense, header_row=3)
            self._apply_data_style(ws_expense, start_row=4)
            self._auto_adjust_columns(ws_expense)
            
            # Add total
            total_expense = sum([row[1] for row in expense_data])
            last_row = len(expense_data) + 4
            ws_expense[f'A{last_row}'] = "TOTAL EXPENSES"
            ws_expense[f'A{last_row}'].font = Font(bold=True)
            ws_expense[f'B{last_row}'] = total_expense
            ws_expense[f'B{last_row}'].font = Font(bold=True)
            
            # Summary sheet
            ws_summary = wb.create_sheet("Summary", 0)
            
            ws_summary['A1'] = "FINANCIAL SUMMARY"
            ws_summary['A1'].font = Font(bold=True, size=16, color="FFFFFF")
            ws_summary['A1'].fill = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
            ws_summary.merge_cells('A1:B1')
            
            ws_summary['A3'] = "Total Income (GHS):"
            ws_summary['B3'] = total_income
            ws_summary['A3'].font = Font(bold=True)
            ws_summary['B3'].font = Font(bold=True, color="27ae60")
            
            ws_summary['A4'] = "Total Expenses (GHS):"
            ws_summary['B4'] = total_expense
            ws_summary['A4'].font = Font(bold=True)
            ws_summary['B4'].font = Font(bold=True, color="e74c3c")
            
            ws_summary['A5'] = "Net Balance (GHS):"
            ws_summary['B5'] = total_income - total_expense
            ws_summary['A5'].font = Font(bold=True)
            ws_summary['B5'].font = Font(bold=True)
            
            # Save file
            if not output_path:
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                output_path = f'financial_summary_{timestamp}.xlsx'
            
            wb.save(output_path)
            self.disconnect()
            
            return True, f"✅ Exported financial summary to {output_path}"
            
        except Exception as e:
            self.disconnect()
            return False, f"❌ Error exporting financial summary: {str(e)}"


def test_export():
    """Test the Excel export functionality"""
    if not OPENPYXL_AVAILABLE:
        print("⚠️  openpyxl not installed. Install with: pip install openpyxl")
        return
    
    exporter = ExcelExporter('school_management.db')
    
    # Test 1: Export all students
    success, message = exporter.export_students_full()
    print(f"Test 1 - Full Student Export: {message}")
    
    # Test 2: Export by class
    success, message = exporter.export_students_by_class()
    print(f"Test 2 - Students by Class: {message}")
    
    # Test 3: Export teacher directory
    success, message = exporter.export_teachers_directory()
    print(f"Test 3 - Teacher Directory: {message}")
    
    # Test 4: Export financial summary
    success, message = exporter.export_financial_summary()
    print(f"Test 4 - Financial Summary: {message}")
    
    # Test 5: Export single student profile (if students exist)
    exporter.connect()
    exporter.cursor.execute('SELECT student_id FROM students LIMIT 1')
    result = exporter.cursor.fetchone()
    exporter.disconnect()
    
    if result:
        student_id = result[0]
        success, message = exporter.export_student_profile(student_id)
        print(f"Test 5 - Student Profile: {message}")
    else:
        print("Test 5 - Student Profile: No students in database")


if __name__ == '__main__':
    test_export()
